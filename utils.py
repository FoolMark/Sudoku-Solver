from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import numpy as np

ROW = np.zeros((9,10),dtype=np.int8)
COL = np.zeros((9,10),dtype=np.int8)
GRID = np.zeros((9,10),dtype=np.int8)

def getMatrix():
    xlsx = load_workbook('data.xlsx')
    in_sheet,_ = xlsx.sheetnames
    assert in_sheet == 'Input Matrix'
    in_sheet = xlsx.worksheets[0]
    data = []
    blank = []
    for i in range(9):
        data.append([0] * 9)
    for r in range(2,11):
        for c in range(2,11):    
            data[r-2][c-2] = in_sheet.cell(r,c).value or 0
            if data[r-2][c-2] == 0:
                blank.append((r-2,c-2))
    return np.array(data),blank

def writeMatrix(origin,solution=None):
    orange_fill = PatternFill(fill_type='solid', fgColor="FFC166")
    blank_fill = PatternFill(fill_type='solid', fgColor="FFFFFF")
    xlsx = load_workbook('data.xlsx')
    _,out_sheet = xlsx.sheetnames
    assert out_sheet == 'Solution Matrix'
    out_sheet = xlsx.worksheets[1]
    for r in range(9):
        for c in range(9):
            out_sheet.cell(r+2,c+2).fill = blank_fill
    for r in range(9):
        for c in range(9):
            if origin[r][c]:
                out_sheet.cell(r+2,c+2,value=origin[r][c])
                out_sheet.cell(r+2,c+2).fill = orange_fill
            else:
                if type(solution) == type(None):
                    continue
                out_sheet.cell(r+2,c+2,value=solution[r][c])
                out_sheet.cell(r+2,c+2).fill = blank_fill

    xlsx.save('data.xlsx')

def rc2g(r,c):
    R = r // 3
    C = c // 3
    g = R * 3 + C
    return g

def checkValid(matrix):
    for i in range(9):
        for j in range(9):
            val = matrix[i][j]
            if val == 0:
                continue
            if ROW[i][val] or COL[j][val] or GRID[rc2g(i,j)][val]:
                return False
            else:
               ROW[i][val] = 1
               COL[j][val] = 1
               GRID[rc2g(i,j)][val] = 1
    return True

if __name__ == '__main__':
    origin,_ = getMatrix()
    checkValid(origin)
    writeMatrix(origin,None)